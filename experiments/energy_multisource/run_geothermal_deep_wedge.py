from __future__ import annotations

import argparse
import hashlib
import json
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def pct_improve(base: float, candidate: float) -> float | None:
    if not np.isfinite(base) or base <= 0:
        return None
    return float(100.0 * (base - candidate) / base)


def forecast_bench(series: pd.Series, horizons: tuple[int, ...], rolling: int) -> dict:
    s = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).ffill(limit=12)
    split = int(len(s) * 0.8)
    out = {}
    for horizon in horizons:
        d = pd.DataFrame(
            {
                "future": s.shift(-horizon),
                "persistence": s,
                "rolling": s.rolling(rolling, min_periods=1).mean(),
            }
        ).iloc[split:].dropna()
        if len(d) < 100:
            continue
        p_mae = float((d["persistence"] - d["future"]).abs().mean())
        r_mae = float((d["rolling"] - d["future"]).abs().mean())
        out[str(horizon)] = {
            "persistence_mae": p_mae,
            "rolling_mae": r_mae,
            "rolling_improvement_pct": pct_improve(p_mae, r_mae),
            "n_test": int(len(d)),
        }
    return out


def analyze_forge1109(flow_path: Path, pressure_path: Path) -> dict:
    flow = pd.read_csv(flow_path)
    pressure = pd.read_csv(pressure_path)
    ft = pd.to_numeric(flow.iloc[:, 0], errors="coerce").to_numpy()
    fq = pd.to_numeric(flow.iloc[:, 1], errors="coerce").to_numpy()
    pt = pd.to_numeric(pressure.iloc[:, 0], errors="coerce").to_numpy()
    pp = pd.to_numeric(pressure.iloc[:, 1], errors="coerce").to_numpy()

    overlap = np.isfinite(pt) & (pt >= np.nanmin(ft)) & (pt <= np.nanmax(ft))
    t = pt[overlap]
    p = pp[overlap]
    q = np.interp(t, ft[np.isfinite(ft) & np.isfinite(fq)], fq[np.isfinite(ft) & np.isfinite(fq)])
    df = pd.DataFrame({"time_hr": t, "pressure_psi": p, "flow_bpm": q}).dropna()
    active = df[df["flow_bpm"] >= 0.5].copy()

    corr = float(active[["pressure_psi", "flow_bpm"]].corr().iloc[0, 1]) if len(active) else None
    active["flow_bin"] = np.round(active["flow_bpm"] * 2.0) / 2.0
    plateaus = (
        active.groupby("flow_bin", observed=True)
        .agg(n=("pressure_psi", "size"), median_pressure_psi=("pressure_psi", "median"), median_flow_bpm=("flow_bpm", "median"))
        .reset_index(drop=True)
    )
    plateaus = plateaus[plateaus["n"] >= 30]

    forecasts = {
        "pressure": forecast_bench(df["pressure_psi"], (12, 60, 180), 12),
        "flow": forecast_bench(df["flow_bpm"], (12, 60, 180), 12),
    }
    forecast_wedges = []
    for target, rows in forecasts.items():
        for horizon, metric in rows.items():
            improve = metric["rolling_improvement_pct"]
            forecast_wedges.append(
                {
                    "source": "DOE_GDR_FORGE_1109",
                    "type": "offline_forecast_screen",
                    "target": target,
                    "horizon_samples": int(horizon),
                    "improvement_pct": improve,
                    "n_test": metric["n_test"],
                    "status": "PROMOTE_FOR_FURTHER_TEST" if improve is not None and improve >= 5 else "HOLD_OR_NEGATIVE",
                }
            )

    return {
        "source": "DOE_GDR_FORGE_1109",
        "flow_sha256": sha256_file(flow_path),
        "pressure_sha256": sha256_file(pressure_path),
        "flow_rows": int(len(flow)),
        "pressure_rows": int(len(pressure)),
        "aligned_overlap_rows": int(len(df)),
        "active_overlap_rows": int(len(active)),
        "pressure_flow_correlation_active": corr,
        "stable_flow_plateaus": plateaus.to_dict(orient="records"),
        "forecast_benchmarks": forecasts,
        "wedges": forecast_wedges,
        "constraint_note": "Pressure and flow are aligned only for an offline historical screen; no safe operating limit, control rule, or causal reservoir parameter is inferred.",
    }


def cycle_table(ws) -> tuple[pd.DataFrame, bool] | None:
    top_text = " ".join(str(v) for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True) for v in row if v is not None).lower()
    unsynced = "not synch" in top_text or "not synchron" in top_text
    headers = [cell.value for cell in ws[3]] if ws.max_row >= 3 else []
    spp_idx = next((i for i, h in enumerate(headers) if isinstance(h, str) and "SPP" in h), None)
    total_idx = next((i for i, h in enumerate(headers) if isinstance(h, str) and "Total Pump Rate" in h), None)
    p3_idx = next((i for i, h in enumerate(headers) if isinstance(h, str) and "Pump 3 Rate" in h), None)
    p2_idx = next((i for i, h in enumerate(headers) if isinstance(h, str) and "Pump 2 Rate" in h), None)
    if spp_idx is None or (total_idx is None and p3_idx is None):
        return None

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            pressure = float(row[spp_idx]) if row[spp_idx] is not None else np.nan
        except Exception:
            pressure = np.nan
        if total_idx is not None:
            try:
                flow = float(row[total_idx]) if row[total_idx] is not None else np.nan
            except Exception:
                flow = np.nan
        else:
            candidates = []
            for idx in (p3_idx, p2_idx):
                if idx is not None:
                    try:
                        candidates.append(float(row[idx]) if row[idx] is not None else np.nan)
                    except Exception:
                        candidates.append(np.nan)
            flow = next((v for v in candidates if np.isfinite(v)), np.nan)
        rows.append((pressure, flow))

    frame = pd.DataFrame(rows, columns=["pressure", "flow"]).dropna()
    if len(frame) < 20:
        return None
    return frame, unsynced


def cycle_metrics(frame: pd.DataFrame) -> dict | None:
    active = frame[frame["flow"] > 5].copy()
    if len(active) < 20:
        return None
    active["dflow"] = active["flow"].diff()
    pressure_range = float(active["pressure"].quantile(0.95) - active["pressure"].quantile(0.05))
    corr = None
    if active["flow"].std() > 0 and active["pressure"].std() > 0:
        corr = float(active[["flow", "pressure"]].corr().iloc[0, 1])

    hysteresis_bins = []
    try:
        bins = pd.qcut(active["flow"], q=min(8, int(active["flow"].nunique())), duplicates="drop")
        active = active.assign(flow_bin=bins)
        for _, group in active.groupby("flow_bin", observed=True):
            rising = group[group["dflow"] > 0]["pressure"]
            falling = group[group["dflow"] < 0]["pressure"]
            if len(rising) >= 3 and len(falling) >= 3:
                hysteresis_bins.append(abs(float(rising.median() - falling.median())))
    except Exception:
        pass

    hysteresis = float(np.median(hysteresis_bins)) if hysteresis_bins else None
    normalized = hysteresis / pressure_range if hysteresis is not None and pressure_range > 0 else None
    return {
        "active_rows": int(len(active)),
        "pressure_flow_correlation": corr,
        "normalized_hysteresis_screen": normalized,
        "hysteresis_bin_count": int(len(hysteresis_bins)),
    }


def analyze_forge1149(paths: list[Path]) -> dict:
    workbooks = []
    time_sync_warning = False
    reviewed_cycles = 0
    strong_path_dependence = 0
    for path in paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        cycles = []
        for ws in wb.worksheets:
            if not ws.title.lower().startswith("cycle"):
                continue
            parsed = cycle_table(ws)
            if parsed is None:
                continue
            frame, unsynced = parsed
            time_sync_warning = time_sync_warning or unsynced
            metric = cycle_metrics(frame)
            if metric is None:
                continue
            metric["sheet"] = ws.title
            metric["clock_sync_warning"] = unsynced
            cycles.append(metric)
            reviewed_cycles += 1
            if metric["normalized_hysteresis_screen"] is not None and metric["normalized_hysteresis_screen"] >= 0.10:
                strong_path_dependence += 1
        workbooks.append({"file": path.name, "sha256": sha256_file(path), "cycles": cycles})

    wedges = []
    if time_sync_warning:
        wedges.append(
            {
                "source": "DOE_GDR_FORGE_1149",
                "type": "sensor_time_alignment",
                "status": "PRIORITY_DATA_QUALITY_WEDGE",
                "evidence": "source workbook states times are not synchronized across tools/data",
                "next_test": "offline timestamp reconciliation and uncertainty propagation before cross-sensor modeling",
            }
        )
    if strong_path_dependence:
        wedges.append(
            {
                "source": "DOE_GDR_FORGE_1149",
                "type": "path_dependence_screen",
                "status": "PRIORITY_MODELING_WEDGE",
                "cycles_flagged": int(strong_path_dependence),
                "cycles_reviewed": int(reviewed_cycles),
                "next_test": "offline state-aware models versus memoryless baselines; do not use for physical control",
            }
        )

    return {
        "source": "DOE_GDR_FORGE_1149",
        "workbooks": workbooks,
        "cycles_reviewed": int(reviewed_cycles),
        "time_sync_warning_detected": bool(time_sync_warning),
        "path_dependence_cycles_flagged": int(strong_path_dependence),
        "wedges": wedges,
        "claim_boundary": "Historical stimulation data are analyzed offline only. This output is not an operational procedure, setpoint recommendation, or safety limit.",
    }


def analyze_usgs(table1: Path, table2: Path, table3: Path) -> dict:
    t1 = pd.read_csv(table1)
    t2 = pd.read_csv(table2)
    t3 = pd.read_csv(table3)

    geo_cols = [
        "Chalcedony_degC",
        "Opal_degC",
        "Giggenbach_degC",
        "KMg_degC",
        "Na_K_degC",
        "NaK_13_Ca_degC",
        "NaK_43_Ca_degC",
        "Mg_corr_NaK_43_Ca_degC",
    ]
    geos = t1[geo_cols].apply(pd.to_numeric, errors="coerce")
    counts = geos.notna().sum(axis=1)
    spread = geos.max(axis=1) - geos.min(axis=1)
    spread_valid = spread[counts >= 3].dropna()

    mean_base = pd.to_numeric(t2["Mean_Accessible_Resource_Base_10to18_J"], errors="coerce")
    plusminus = pd.to_numeric(t2["PlusMinus_of_Mean_Accessible_Resource_Base_10to18_J"], errors="coerce")
    relative_uncertainty = (plusminus / mean_base.replace(0, np.nan)).dropna()

    completeness = {}
    for col in ["Temperature_degC", "Flow_L_min", "pH", "TDS_mg_L"] + geo_cols:
        numeric = pd.to_numeric(t1[col], errors="coerce")
        completeness[col] = float(numeric.notna().mean())

    state_counts = t2.groupby("State").size().sort_values(ascending=False).head(15)

    return {
        "source": "USGS_GEOTHERMAL_OFR83250",
        "sha256": {"table1": sha256_file(table1), "table2": sha256_file(table2), "table3": sha256_file(table3)},
        "rows": {"table1": int(len(t1)), "table2": int(len(t2)), "table3": int(len(t3))},
        "data_completeness_fraction": completeness,
        "geothermometer_disagreement_degC": {
            "sites_with_at_least_3_estimators": int(len(spread_valid)),
            "median_spread": float(spread_valid.median()),
            "p90_spread": float(spread_valid.quantile(0.90)),
            "p95_spread": float(spread_valid.quantile(0.95)),
        },
        "accessible_resource_relative_uncertainty": {
            "systems": int(len(relative_uncertainty)),
            "median": float(relative_uncertainty.median()),
            "p90": float(relative_uncertainty.quantile(0.90)),
            "fraction_above_50pct": float((relative_uncertainty > 0.50).mean()),
        },
        "top_states_by_system_count": {str(k): int(v) for k, v in state_counts.items()},
        "wedges": [
            {
                "source": "USGS_GEOTHERMAL_OFR83250",
                "type": "resource_uncertainty_reconciliation",
                "status": "PRIORITY_SCREENING_WEDGE",
                "next_test": "compare point-estimate ranking with uncertainty-aware ranking and hold out sites with sparse chemistry",
            },
            {
                "source": "USGS_GEOTHERMAL_OFR83250",
                "type": "missingness_robustness",
                "status": "PRIORITY_DATA_QUALITY_WEDGE",
                "next_test": "stress-test rankings under chemistry/flow missingness rather than imputing unsupported certainty",
            },
        ],
        "claim_boundary": "Historical USGS resource screening only; not a current reserve estimate, project economics, site recommendation, or bankable resource assessment.",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--forge1109-dir", required=True)
    parser.add_argument("--forge1149-dir", required=True)
    parser.add_argument("--usgs-dir", required=True)
    parser.add_argument("--out", default="out/geothermal-deep-analysis")
    args = parser.parse_args()

    f1109 = Path(args.forge1109_dir)
    f1149 = Path(args.forge1149_dir)
    usgs = Path(args.usgs_dir)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    stimulation_files = [
        f1149 / "58-32_OpenHoleStimulation (1).xlsx",
        f1149 / "58-32_LowerPerfStimulation (1).xlsx",
        f1149 / "58-32_UpperPerforationStimulation_mm (1).xlsx",
    ]
    tables = list(usgs.rglob("OFR83250_Table*.csv"))
    table_map = {p.stem: p for p in tables}

    result = {
        "schema_version": "1.0",
        "negative_results_retained": True,
        "experiments": [
            analyze_forge1109(f1109 / "flow.csv", f1109 / "pressure.csv"),
            analyze_forge1149([p for p in stimulation_files if p.exists()]),
            analyze_usgs(
                table_map["OFR83250_Table1"],
                table_map["OFR83250_Table2"],
                table_map["OFR83250_Table3"],
            ),
        ],
        "claim_boundary": "Offline retrospective research only. No live control, safety, siting, reserve, or commercial-performance recommendation is produced.",
    }
    result["wedges"] = []
    for experiment in result["experiments"]:
        result["wedges"].extend(experiment.get("wedges", []))

    path = out / "summary.json"
    path.write_text(json.dumps(result, indent=2, default=str) + "\n", encoding="utf-8")
    manifest = {path.name: sha256_file(path)}
    (out / "sha256_manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"wedges": result["wedges"]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
