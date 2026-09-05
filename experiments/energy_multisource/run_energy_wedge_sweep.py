from __future__ import annotations

import argparse
import hashlib
import json
import math
from pathlib import Path

import numpy as np
import pandas as pd

RHO_WATER = 1025.0
G = 9.80665


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def finite(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)


def pct_improve(base: float, candidate: float) -> float | None:
    if not np.isfinite(base) or base <= 0:
        return None
    return 100.0 * (base - candidate) / base


def forecast_bench(series: pd.Series, horizons=(20, 60, 120), rolling=20) -> dict:
    s = finite(series).ffill(limit=10)
    out: dict[str, dict] = {}
    split = int(len(s) * 0.8)
    for horizon in horizons:
        future = s.shift(-horizon)
        rolling_candidate = s.rolling(rolling, min_periods=1).mean()
        d = pd.DataFrame(
            {"y": future, "persistence": s, "rolling": rolling_candidate}
        ).iloc[split:].dropna()
        if len(d) < 30:
            continue
        metrics: dict[str, dict | float | int | None] = {}
        for name in ("persistence", "rolling"):
            err = (d[name] - d["y"]).abs()
            metrics[name] = {
                "mae": float(err.mean()),
                "rmse": float(np.sqrt(((d[name] - d["y"]) ** 2).mean())),
            }
        metrics["rolling_improvement_pct"] = pct_improve(
            metrics["persistence"]["mae"], metrics["rolling"]["mae"]
        )
        metrics["n_test"] = int(len(d))
        out[str(horizon)] = metrics
    return out


def load_forge1683(xlsx: Path) -> pd.DataFrame:
    # Excel ingestion is optional for wave analysis and the repository test suite.
    # Keep the requirement local to the only path that reads an XLSX source.
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        if exc.name != "openpyxl":
            raise
        raise RuntimeError(
            "FORGE XLSX ingestion requires the optional openpyxl dependency; "
            "install the research acquisition dependencies before using --forge-xlsx."
        ) from exc
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Data (30 sec)"] if "Data (30 sec)" in wb.sheetnames else wb[wb.sheetnames[0]]
    labels = [
        "time",
        "p16b_psi",
        "flow16b_1_gpm",
        "flow16b_2_gpm",
        "temp16b_f",
        "sep_flow1_gpm",
        "sep_flow2_gpm",
        "sep_total_gpm",
        "p16a_psi",
        "pump_rate_bpm",
        "liberty_whp_psi",
    ]
    rows = []
    for source_row in ws.iter_rows(min_row=5, values_only=True):
        row = [source_row[0]]
        for value in source_row[1:11]:
            try:
                row.append(float(value) if value not in (None, "") else np.nan)
            except Exception:
                row.append(np.nan)
        rows.append(row)
    df = pd.DataFrame(rows, columns=labels)
    df["time"] = pd.to_datetime(df["time"], errors="coerce")
    return df


def lag_corr(a: pd.Series, b: pd.Series, max_lag=240) -> dict | None:
    x = finite(a).interpolate(limit=10).diff().to_numpy()
    y = finite(b).interpolate(limit=10).diff().to_numpy()
    best = None
    for lag in range(max_lag + 1):
        xa = x[:-lag] if lag else x
        yb = y[lag:] if lag else y
        mask = np.isfinite(xa) & np.isfinite(yb)
        if mask.sum() < 1000:
            continue
        corr = float(np.corrcoef(xa[mask], yb[mask])[0, 1])
        candidate = {
            "lag_samples": lag,
            "lag_minutes": lag * 0.5,
            "corr": corr,
            "n": int(mask.sum()),
        }
        if best is None or abs(corr) > abs(best["corr"]):
            best = candidate
    return best


def analyze_forge1683(xlsx: Path) -> dict:
    df = load_forge1683(xlsx)
    fields = [c for c in df.columns if c != "time"]
    quality = {}
    for col in fields:
        values = finite(df[col])
        valid = values.dropna()
        quality[col] = {
            "valid": int(valid.size),
            "missing_pct": round(float(values.isna().mean() * 100), 3),
            "min": float(valid.min()) if len(valid) else None,
            "median": float(valid.median()) if len(valid) else None,
            "p95": float(valid.quantile(0.95)) if len(valid) else None,
            "max": float(valid.max()) if len(valid) else None,
        }

    active = df[(finite(df["pump_rate_bpm"]) > 5) & (finite(df["sep_total_gpm"]) > 10)].copy()
    active["inj_gpm"] = active["pump_rate_bpm"] * 42.0
    active["recovery_ratio"] = active["sep_total_gpm"] / active["inj_gpm"]
    recovery = finite(active["recovery_ratio"]).dropna()
    q05 = float(recovery.quantile(0.05))
    q95 = float(recovery.quantile(0.95))
    low = active[active["recovery_ratio"] <= q05]
    high = active[active["recovery_ratio"] >= q95]

    edge_compare = {}
    for col in [
        "pump_rate_bpm",
        "sep_total_gpm",
        "p16a_psi",
        "p16b_psi",
        "temp16b_f",
        "liberty_whp_psi",
    ]:
        edge_compare[col] = {
            "low_recovery_median": float(finite(low[col]).median()),
            "high_recovery_median": float(finite(high[col]).median()),
        }

    forecasts = {
        "sep_total_gpm": forecast_bench(active["sep_total_gpm"]),
        "p16b_psi": forecast_bench(active["p16b_psi"]),
        "temp16b_f": forecast_bench(active["temp16b_f"]),
    }

    wedges = []
    for target, results in forecasts.items():
        for horizon, metrics in results.items():
            improvement = metrics.get("rolling_improvement_pct")
            status = (
                "PROMOTE_FOR_FURTHER_TEST"
                if improvement is not None and improvement >= 5 and metrics["n_test"] >= 1000
                else "HOLD_OR_NEGATIVE"
            )
            wedges.append(
                {
                    "source": "FORGE1683",
                    "target": target,
                    "horizon_minutes": int(horizon) * 0.5,
                    "candidate": "rolling20",
                    "improvement_pct": improvement,
                    "n_test": metrics["n_test"],
                    "status": status,
                }
            )

    return {
        "source": "DOE_GDR_FORGE_1683",
        "source_sha256": sha256_file(xlsx),
        "samples": int(len(df)),
        "start": str(df["time"].min()),
        "end": str(df["time"].max()),
        "quality": quality,
        "active_circulation_rows": int(len(active)),
        "recovery_ratio_proxy": {
            "definition": "separator total flow / (pump rate * 42 gal/bbl); contemporaneous screening proxy, not reservoir recovery accounting",
            "median": float(recovery.median()),
            "mean": float(recovery.mean()),
            "p05": q05,
            "p95": q95,
        },
        "edge_case_compare": edge_compare,
        "lag_screening": {
            "pump_to_sep_total_first_difference": lag_corr(
                active["pump_rate_bpm"], active["sep_total_gpm"]
            ),
            "p16a_to_p16b_first_difference": lag_corr(active["p16a_psi"], active["p16b_psi"]),
        },
        "forecast_benchmarks": forecasts,
        "wedges": wedges,
        "claim_boundary": "Field data are external, but analyses are internal screening. No causal, operational, safety, or commercial-performance claim is authorized.",
    }


def analyze_ndbc_file(path: Path) -> dict | None:
    data = []
    for line in path.read_text(errors="ignore").splitlines():
        if not line.strip() or line.startswith("#"):
            continue
        parts = line.split()
        if len(parts) < 16:
            continue
        try:
            year, month, day, hour, minute = map(int, parts[:5])
            values = []
            for raw in parts[5:]:
                try:
                    values.append(float(raw) if raw not in {"MM", "999", "9999", "99.0"} else np.nan)
                except Exception:
                    values.append(np.nan)
            data.append(
                [pd.Timestamp(year=year, month=month, day=day, hour=hour, minute=minute)]
                + values
            )
        except Exception:
            continue

    names = [
        "time",
        "WDIR",
        "WSPD",
        "GST",
        "WVHT",
        "DPD",
        "APD",
        "MWD",
        "PRES",
        "ATMP",
        "WTMP",
        "DEWP",
        "VIS",
        "PTDY",
        "TIDE",
    ]
    if not data:
        return None
    width = min(len(names), min(len(row) for row in data))
    df = pd.DataFrame([row[:width] for row in data], columns=names[:width]).sort_values("time")
    if not {"WVHT", "APD"}.issubset(df.columns):
        return None

    df["wave_power_proxy_kw_m"] = (
        (RHO_WATER * G * G / (64 * math.pi))
        * finite(df["WVHT"]) ** 2
        * finite(df["APD"])
        / 1000.0
    )
    valid = df.dropna(subset=["wave_power_proxy_kw_m"]).reset_index(drop=True)
    if len(valid) < 30:
        return None

    dt_minutes = valid["time"].diff().dropna().dt.total_seconds() / 60
    median_step = float(dt_minutes.median()) if len(dt_minutes) else None
    if median_step and median_step > 0:
        horizons = tuple(max(1, int(round(minutes / median_step))) for minutes in (60, 180, 360))
    else:
        horizons = (1, 3, 6)

    power = valid["wave_power_proxy_kw_m"]
    forecast = forecast_bench(power, horizons=horizons, rolling=max(2, horizons[0]))
    ramp = power.diff().abs()

    return {
        "file": path.name,
        "samples": int(len(valid)),
        "median_step_minutes": median_step,
        "wave_power_proxy_kw_m": {
            "median": float(power.median()),
            "p95": float(power.quantile(0.95)),
            "max": float(power.max()),
            "formula": "rho*g^2/(64*pi)*Hs^2*APD; deep-water screening proxy using APD",
        },
        "ramp_threshold_p90_kw_m": float(ramp.quantile(0.90)),
        "forecast_benchmarks": forecast,
        "claim_boundary": "Resource proxy only; not device power, capacity factor, LCOE, or site-bankable yield.",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--forge-xlsx")
    parser.add_argument("--wave-dir")
    parser.add_argument("--out", default="out/energy-multisource")
    args = parser.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    result = {
        "schema_version": "1.0",
        "experiments": [],
        "wedges": [],
        "negative_results_retained": True,
    }

    if args.forge_xlsx and Path(args.forge_xlsx).exists():
        geothermal = analyze_forge1683(Path(args.forge_xlsx))
        result["experiments"].append(geothermal)
        result["wedges"].extend(geothermal["wedges"])

    if args.wave_dir and Path(args.wave_dir).exists():
        waves = []
        for path in sorted(Path(args.wave_dir).glob("*.txt")):
            item = analyze_ndbc_file(path)
            if item:
                waves.append(item)
        result["experiments"].append(
            {"source": "NOAA_NDBC_WAVES", "stations": waves, "station_count": len(waves)}
        )
        for wave in waves:
            for horizon, metrics in wave["forecast_benchmarks"].items():
                improvement = metrics.get("rolling_improvement_pct")
                status = (
                    "PROMOTE_FOR_FURTHER_TEST"
                    if improvement is not None and improvement >= 5 and metrics["n_test"] >= 30
                    else "HOLD_OR_NEGATIVE"
                )
                result["wedges"].append(
                    {
                        "source": "NOAA_NDBC_WAVES",
                        "station": wave["file"],
                        "target": "wave_power_proxy",
                        "horizon_samples": int(horizon),
                        "improvement_pct": improvement,
                        "n_test": metrics["n_test"],
                        "status": status,
                    }
                )

    result["wedges"] = sorted(
        result["wedges"],
        key=lambda row: (
            row.get("improvement_pct") is not None,
            row.get("improvement_pct") if row.get("improvement_pct") is not None else -1,
        ),
        reverse=True,
    )

    summary = out / "summary.json"
    summary.write_text(json.dumps(result, indent=2, default=str) + "\n", encoding="utf-8")
    manifest = {
        path.name: sha256_file(path)
        for path in out.iterdir()
        if path.is_file() and path.name != "sha256_manifest.json"
    }
    (out / "sha256_manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps({"out": str(out), "top_wedges": result["wedges"][:10]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
